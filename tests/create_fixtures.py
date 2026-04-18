"""Script to create test fixture files that require binary formats."""
from pathlib import Path
import openpyxl
from fpdf import FPDF

fixtures = Path(__file__).parent / "fixtures"
fixtures.mkdir(exist_ok=True)

wb = openpyxl.Workbook()

# Sheet 1: Employees
ws1 = wb.active
ws1.title = "Employees"
ws1.append(["id", "name", "department", "salary"])
ws1.append([1, "Alice Johnson", "Engineering", 95000])
ws1.append([2, "Bob Smith", "Marketing", 72000])
ws1.append([3, "Carol White", "Engineering", 105000])

# Sheet 2: Projects
ws2 = wb.create_sheet("Projects")
ws2.append(["project_id", "name", "status", "owner"])
ws2.append([101, "Phoenix", "active", "Alice Johnson"])
ws2.append([102, "Atlas", "planning", "Carol White"])

# Sheet 3: Empty (should be skipped)
ws3 = wb.create_sheet("Empty")

wb.save(fixtures / "sample.xlsx")
print(f"Created {fixtures / 'sample.xlsx'}")

# ── sample.pdf ────────────────────────────────────────────────────────────────
# Simulates a short white paper with numbered sections and a table.
pdf = FPDF()
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
pdf.set_font("Helvetica", "B", 16)
pdf.cell(0, 10, "LLM Architecture White Paper", new_x="LMARGIN", new_y="NEXT")
pdf.set_font("Helvetica", "", 11)
pdf.ln(4)

sections = [
    ("ABSTRACT",
     "This paper examines large language model architectures and their applications in "
     "knowledge management systems. We evaluate transformer-based models for document "
     "ingestion, semantic search, and automated wiki generation."),
    ("1. INTRODUCTION",
     "Large language models (LLMs) have transformed natural language processing. "
     "Recent advances in models such as GPT-4 and Claude have demonstrated remarkable "
     "capabilities in text generation, summarisation, and reasoning. "
     "This paper explores their use in auto-evolving documentation systems."),
    ("2. METHODOLOGY",
     "We use a pipeline consisting of (a) document parsing, (b) LLM-based page generation, "
     "(c) fidelity validation, (d) semantic consistency checks, and (e) vector-based "
     "cross-linking using sentence-transformer embeddings stored in ChromaDB."),
    ("3. RESULTS",
     "Experiments on 50 enterprise data files show 94% fidelity scores on first pass, "
     "with 99% after one regeneration cycle. Consistency violations were detected in "
     "3% of cases, preventing contradictory pages from entering the wiki."),
    ("4. CONCLUSION",
     "Automated wiki generation using LLMs is both feasible and accurate when combined "
     "with structured validation pipelines. Future work will extend support to "
     "image-heavy documents via OCR integration."),
]

for heading, body in sections:
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, heading, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(0, 6, body)
    pdf.ln(4)

pdf.output(fixtures / "sample.pdf")
print(f"Created {fixtures / 'sample.pdf'}")
