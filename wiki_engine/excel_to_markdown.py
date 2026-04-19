from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl


def _slugify(text: str, fallback: str = "entry") -> str:
    text = text.strip().lower().replace(" ", "-")
    text = re.sub(r"[^a-z0-9\-]", "", text)
    return text or fallback


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _read_sheet_data(sheet) -> tuple[list[str], list[list[str]]]:
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    non_empty = [
        row for row in rows
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]
    if not non_empty:
        return [], []
    header = [str(cell).strip() if cell is not None else "" for cell in non_empty[0]]
    data_rows = [[_format_value(cell) for cell in row] for row in non_empty[1:]]
    return header, data_rows


def _format_row_markdown(
    row: dict[str, str],
    workbook_name: str,
    sheet_name: str,
    row_number: int,
) -> str:
    identifier = str(row[list(row.keys())[0]])
    title = identifier

    tags = [
        "excel",
        _slugify(workbook_name, "workbook"),
        _slugify(sheet_name, "sheet"),
        _slugify(identifier),
    ]

    name_field = row.get("name", identifier)
    summary = f"This entry describes {name_field}."

    data_points = [f"- **{key}**: {value}" for key, value in row.items()]

    content_lines = [
        "---",
        f"title: {title}",
        f"tags: [{', '.join(tags)}]",
        f"sources: [\"{workbook_name}\"]",
        f"sheet: {sheet_name}",
        f"row number: {row_number}",
        "---",
        "",
        "## Summary",
        "",
        summary,
        "",
        "## Data Points",
        "",
        "\n".join(data_points),
    ]

    return "\n".join(content_lines).strip() + "\n"


def dump_excel_to_markdown(
    file_path: Path,
    wiki_dir: Path,
    sheet_name: str | None = None,
) -> list[Path]:
    """Read an Excel workbook and write one markdown page per data row.

    Returns the list of written file paths.
    """
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    workbook_name = file_path.stem
    workbook_slug = _slugify(workbook_name, "workbook")

    output_paths: list[Path] = []

    for name in workbook.sheetnames:
        if sheet_name is not None and name != sheet_name:
            continue

        header, data_rows = _read_sheet_data(workbook[name])
        if not header:
            continue

        sheet_slug = _slugify(name, "sheet")
        output_dir = wiki_dir / "excel" / workbook_slug / sheet_slug
        output_dir.mkdir(parents=True, exist_ok=True)

        for row_num, row_values in enumerate(data_rows, start=1):
            if not any(v.strip() for v in row_values):
                continue

            row = dict(zip(header, row_values))
            identifier = row_values[0] if row_values else f"row-{row_num}"
            slug = _slugify(identifier) or f"row-{row_num}"

            # Ensure unique filenames when slugs collide
            candidate = output_dir / f"{slug}.md"
            if candidate.exists():
                slug = f"{slug}-{row_num}"
                candidate = output_dir / f"{slug}.md"

            content = _format_row_markdown(row, workbook_name, name, row_num)
            candidate.write_text(content, encoding="utf-8")
            output_paths.append(candidate)

    workbook.close()
    return output_paths
