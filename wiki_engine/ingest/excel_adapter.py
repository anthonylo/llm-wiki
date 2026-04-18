from pathlib import Path

import openpyxl

from .base import IngestAdapter, SourceSection

MAX_ROWS = 200


class ExcelAdapter(IngestAdapter):
    """Adapter for .xlsx/.xls files — one SourceSection per sheet."""

    def parse(self) -> list[SourceSection]:
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        sections: list[SourceSection] = []
        sheet_slugs: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Skip empty sheets
            non_empty = [r for r in rows if any(c is not None for c in r)]
            if not non_empty:
                continue

            header = [str(c) if c is not None else "" for c in non_empty[0]]
            data_rows = non_empty[1:]
            total_rows = len(data_rows)
            capped = data_rows[:MAX_ROWS]

            lines = ["\t".join(header)]
            for row in capped:
                lines.append("\t".join(str(c) if c is not None else "" for c in row))

            note = ""
            if total_rows > MAX_ROWS:
                note = f"\n\n[Note: Showing first {MAX_ROWS} of {total_rows} rows.]"

            raw_text = "\n".join(lines) + note

            section = SourceSection(
                source_file=self.file_path,
                section_name=sheet_name,
                content_type="excel",
                raw_text=raw_text,
                row_count=total_rows,
                column_names=header,
                metadata={"workbook": self.file_path.stem},
            )
            sections.append(section)
            sheet_slugs.append(section.slug)

        wb.close()

        if not sections:
            return []

        # Generate workbook index section
        index_lines = [f"# {self.file_path.stem} — Workbook Index", ""]
        index_lines.append("This workbook contains the following sheets:")
        index_lines.append("")
        for sec, slug in zip(sections, sheet_slugs):
            index_lines.append(f"- [{sec.section_name}]({slug}.md)")

        index_section = SourceSection(
            source_file=self.file_path,
            section_name="index",
            content_type="excel",
            raw_text="\n".join(index_lines),
            row_count=0,
            column_names=[],
            metadata={"workbook": self.file_path.stem, "is_index": True, "sheet_slugs": sheet_slugs},
        )
        sections.append(index_section)
        return sections
